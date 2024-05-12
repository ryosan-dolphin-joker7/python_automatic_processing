import openpyxl as excel
import datetime

# 新規ワークブックを作りワークシートを得る
book = excel.Workbook()
sheet = book.active

# 今年を得る --- (*1)
thisyear = datetime.date.today().year

# 連続でセルに値を設定する --- (*2)
for i in range(80):
    # 設定する値を計算
    age = i # 満年齢
    year = thisyear - i # 生年
    # セルを取得して値を設定 --- (*3)
    age_cell = sheet.cell(i+1, 1)
    age_cell.value = str(i) + "才"
    year_cell = sheet.cell(i+1, 2)
    year_cell.value = str(year) + "年"

# ファイルを保存
book.save("agelist.xlsx")

