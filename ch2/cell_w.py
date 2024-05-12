# openpyxlを取り込む
import openpyxl as excel

# ワークブックを生成しワークシートを得る
book = excel.Workbook()
sheet = book.active

# A1に値を書き込む --- (*1)
sheet["A1"] = "勤勉な人の計画は必ず成功する"

# A2(row=2, column=1)に値を書き込む --- (*2)
sheet.cell(row=2, column=1, value="猿の尻笑い")

# A3(row=3, column=1)に値を書き込む --- (*3)
cell = sheet.cell(row=3, column=1)
cell.value = "捜すのに時があり諦めるのに時がある"

# 保存
book.save("cell_w.xlsx")
