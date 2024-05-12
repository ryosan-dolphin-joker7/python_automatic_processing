import openpyxl as excel
import random

# 当たりシートの番号を決める --- (*1)
atari = random.randint(1,100)

# 新規ブックを作成 --- (*2)
book = excel.Workbook()
book.active["B2"] = "当たりが書かれたシートを探そう"

# 繰り返し100回シートを作成する --- (*3)
for i in range(1,101):
    # 新規シート作成 --- (*4)
    sname = str(i) + "番"
    sheet = book.create_sheet(title=sname)
    # シートに書き込む単語を決定
    word = "ハズレ"
    if i == atari: word = "当たり"
    # インパクトがあるようにwordで画面を埋める --- (*5)
    for y in range(50):
        for x in range(30):
            c = sheet.cell(y+1, x+1)
            c.value = word

# ファイルに保存 --- (*6)
book.save("game100.xlsx")
print("ok, atari=", atari)

