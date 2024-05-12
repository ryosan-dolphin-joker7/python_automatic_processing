import openpyxl as excel
import datetime

# 新規ワークブックを作りワークシートを得る
book = excel.Workbook()
sheet = book.active

# 開始年を設定 --- (*1)
base_year = datetime.date.today().year - 10

# 連続でセルに値を設定する --- (*2)
for i in range(50):
    # 設定する値を計算 --- (*3)
    year = base_year + i
    s1 = year - 7 # 4/2以降に生まれた人
    s2 = year - 6 # 早生まれの人
    sf = "{}年4/2〜{}年4/1に生まれた人".format(s1, s2)
    # セルに値を設定 --- (*4)
    sheet.cell(i+1, 1, value=str(year)+"年度")
    sheet.cell(i+1, 2, value=sf)

# ファイルを保存
book.save("nyugaku_list.xlsx")

