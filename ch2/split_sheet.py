import openpyxl as excel

# 顧客一覧のブックを開いて顧客一覧ページを得る --- (*1)
book = excel.load_workbook("all-customer.xlsx")
sheet = book["名簿"]

# 顧客一覧を確認してシートに分ける --- (*2)
for row in sheet.iter_rows(min_row=3):
    cells = [v.value for v in row]
    if cells[0] is None: break
    print(cells)
    # 各列のデータを変数に代入 --- (*3)
    (name,area,plan) = cells
    # コピー先のシート名を決める --- (*4)
    sname = plan+"プラン"
    # 該当するシートがあるか --- (*5)
    if sname not in book.sheetnames:
        to_sheet = book.create_sheet(title=sname)
        to_sheet.append(["名前","住所","プラン"])
    else:
        to_sheet = book[sname]
    # 該当シートに顧客情報を追記 --- (*6)
    to_sheet.append(cells)

# ファイルに書き込む --- (*7)
book.save("split_sheet.xlsx")
