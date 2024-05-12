import openpyxl as excel

# 売上データのブックを開いてシートを取り出す
book = excel.load_workbook("uriage-keisen.xlsx")
sheet = book.active
print("最下行", sheet.max_row)
print("最右列", sheet.max_column)
