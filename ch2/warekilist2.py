import openpyxl as excel

# 新規ワークブックを作ってシートを得る
book = excel.Workbook()
sheet = book.active

# ワークシートのヘッダ部分に説明を入れる --- (*1)
sheet["A1"] = "西暦"
sheet["B1"] = "和暦"

# 100年分の西暦和暦の対応表を作る --- (*2)
start_y = 1930
for i in range(100):
    # 西暦と和暦への計算式を設定 --- (*3)
    sei = str(start_y + i)
    wa = '=TEXT("{}/1/1","ggge年")'.format(sei)
    # ワークシートに設定
    sheet.cell(row=(2+i), column=1, value=sei+'年')
    sheet.cell(row=(2+i), column=2, value=wa)
    print(sei, "=", wa)

# ファイルを保存
book.save("wareki2.xlsx")