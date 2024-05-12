import openpyxl as excel
from openpyxl.styles import PatternFill
import datetime

in_file = 'date-sample.xlsx'
out_file = 'date-mark_olditems.xlsx'
limit = datetime.datetime(2020, 1, 1)

# メイン処理
def date_mark(in_file, out_file):
    # Excelブックを開く --- (*1)
    book = excel.load_workbook(in_file)
    # 全シートを確認する --- (*2)
    for sheet in book.worksheets:
        # シート内の範囲を調べる --(*3)
        for row in sheet.iter_rows(min_row=4):
            check_row(row)
   # 保存
    book.save(out_file)

# 指定行をチェックする --- (*4)
def check_row(row):
    date = row[0].value # A列を得る --- (*5)
    # セルが日付型でなければチェックしない
    if type(date) is not datetime.datetime:
        return
    # 古い在庫か？ --- (*6)
    if date < limit:
        # 背景を赤色に --- (*7)
        red = PatternFill(
            fill_type='solid', 
            fgColor='ff0000')
        # その行のセルを全て赤色に設定 --- (*8)
        for cell in row:
            cell.fill = red

if __name__ == '__main__':
    date_mark(in_file, out_file)

