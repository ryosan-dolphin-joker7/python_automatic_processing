import openpyxl as excel, json
# 各種設定 --- (*1)
in_file = 'matome.json'
out_dir = './invoice02'
template_file = 'invoice-template.xlsx'
subject = "2月分のご請求"

#メイン処理 --- (*2)
def gen_invoice():
    # JSONファイルを読み込む --- (*3)
    with open(in_file, "rt") as fp:
        users = json.load(fp)
    # 顧客ごとに請求書を作成 --- (*4)
    for name, data in users.items():
        make_user_invoice(name, data)

# テンプレートにデータを流し込む --- (*5)
def make_user_invoice(name, data):
    # テンプレートを読み込む --- (*6)
    book = excel.load_workbook(template_file)
    sheet = book.active
    # シートに名前と件名と金額を書き込む --- (*7)
    sheet["B4"] = name
    sheet["C10"] = subject
    sheet["C11"] = data["total"]
    # 内訳を連続で書き込む --- (*8)
    for i, it in enumerate(data['items']):
        date, summary, cnt, price = it
        row = 15 + i
        sheet.cell(row, 2, summary+'('+date+')')
        sheet.cell(row, 5, cnt)
        sheet.cell(row, 6, price)
        sheet.cell(row, 7, cnt*price)
    # 請求書をファイルに保存 --- (*9)
    out_file = out_dir+'/' + name + '様.xlsx'
    book.save(out_file)
    print("save: ", out_file)

if __name__ == "__main__":
    gen_invoice() # メイン処理を実行
