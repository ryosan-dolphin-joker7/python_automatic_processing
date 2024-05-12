import openpyxl as excel
# 入出力ファイルを指定
in_file = 'name2.xlsx'
out_file = 'name_combine.xlsx'

# 入力ブックを開きシートを得る --- (*1)
in_book = excel.load_workbook(in_file)
in_sheet = in_book.worksheets[0]
# 新規ブックを生成しシートを得る --- (*2)
out_book = excel.Workbook()
out_sheet = out_book.active
# シートを全部読む --- (*3)
for row in in_sheet.iter_rows():
    # 行内の姓と名のデータを得る --- (*4)
    sei = row[0].value
    na = row[1].value
    # 姓と名をつなげる --- (*5)
    name = sei + ' ' + na
    # 新規シートに追記 --- (*6)
    out_sheet.append([name])
# 結果を保存
out_book.save(out_file)

