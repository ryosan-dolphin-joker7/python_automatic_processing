import openpyxl as excel
# 入出力ファイルを指定
in_file = 'name1.xlsx'
out_file = 'name_split.xlsx'

# 入力ブックを開いてシートを得る --- (*1)
book = excel.load_workbook(in_file)
sheet = book.worksheets[0]
# 新規ブックを作成しシートを得る --- (*2)
out_book = excel.Workbook()
out_sheet = out_book.active
# シートを全部読む --- (*3)
for row in sheet.iter_rows():
    # 名前のセルを得る --- (*4)
    name = row[0].value
    # 姓と名に分ける --- (*5)
    sei, na = name.split(' ')
    # 新規シートに追加 --- (*6)
    out_sheet.append([sei, na])
# 結果を保存
out_book.save(out_file)

